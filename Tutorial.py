import json
import itertools
import pprint
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

grades_book = load_workbook('Grades.xlsx')
new_grades_book = load_workbook('New Grades.xlsx')
ws = grades_book.active
ws2 = new_grades_book.active


def get_headers(ws):
    value = []
    row = [i[0] for i in ws.iter_cols(values_only=True)]

    # Deconstruct list
    for i in row:
        value += [i]

    return value

def populate_col(sheet, total_columns):
    print('NEEDS TO BE COMPLETED')

def get_column(column, sheet, row) -> list:
    col = []

    for i in sheet.iter_rows(min_row = row, values_only = True):
        col.append(i[column])

    return col

def insert_every_n(l1, l2, k):
    i1, i2 = iter(l1), iter(l2)
    while True:
        try:
            yield from itertools.islice(i1, k)
            yield next(i2)
        except StopIteration:
            return

# create the list of headings for amount of rows in the column
heading = get_column(1, ws, 1)[0]
head = [heading] * len(get_column(1, ws, 2))

# create list of names
name = get_column(1, ws, 2)
new_list = list(insert_every_n(head, name, k=1)) # combine the list of names and headings

results = []
values = {}
for key, value in zip(new_list[0::2], new_list[1::2]):
    values = dict(zip([key], [value]))
    results.append(values)

data = list(insert_every_n(get_column(0, ws, 2), results, k=1))

def convert(lst):
    result = {lst[i]: lst[i + 1] for i in range(0, len(lst), 2)}
    return result

# pprint.pprint(convert(data))

new_data = convert(data)
headings = ['ID'] + list(data[1].keys())
# ws2.append(headings)
# print(convert(data))
for person in new_data:
	info = list(new_data[person].values())
	ws2.append([person] + name)
    # print([person] + name)
new_grades_book.save('New Grades.xlsx')

data2 = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}